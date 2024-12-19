import sys
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QLabel, QLineEdit,QFormLayout, QPushButton, QVBoxLayout,QDateEdit,QSpinBox,QComboBox,QFileDialog,QTableWidget,QHeaderView,
    QHBoxLayout, QFrame, QMessageBox, QWidget,QApplication, QWidget, QVBoxLayout, QTableWidget, QTableWidgetItem, QPushButton, QHeaderView,QGridLayout
) 
from PyQt5.QtGui import QPixmap, QPalette, QBrush
from PyQt5.QtCore import Qt, QDate
import sqlite3 
from PyQt5.QtGui import QColor

import sys 
import pandas as pd
from PyQt5.QtWidgets import QApplication, QVBoxLayout, QHBoxLayout, QWidget, QTableWidget, QTableWidgetItem
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from datetime import datetime
from openpyxl import Workbook
import os
import shutil
from io import BytesIO
from tempfile import NamedTemporaryFile
from reportlab.lib.pagesizes import letter
import pandas as pd
import matplotlib.pyplot as plt
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader

import numpy as np
import hashlib
 
def generer_hash_md5(message):
    hash_object = hashlib.md5(message.encode())
    return hash_object.hexdigest()
 
def verifier_hash_md5(message, hash_md5): 
    hash_calcule = generer_hash_md5(message)
    return hash_calcule == hash_md5






from pathlib import Path
import hashlib
from datetime import datetime, timedelta 

current_directory = Path(__file__).parent
racine = current_directory.parent 



dataset = Path.home()/"dataset"


path_profils_images = Path.home()/"images"/"profiles"

path_data_set = dataset/"royal_fitness.db"


logo_path = racine/"images"/"logos"/"logoa.png"


background_path = racine / "style"/"image.jpg"
path_code = racine /"dataset"/"text.txt"
arrowdrop = racine/"images"/"logos"/"ic_arrow_drop_down_black_18dp_1x.png"

from dataset.dataset import *




def code_f(code):
    list_ = code.split('-') 
    new_date = datetime.now()    
    for i in range(0,7):
        date_i = new_date - timedelta(days=i)
        
        code1=f"{list_[0]}-{generer_hash_md5(generer_hash_md5(date_i.strftime('%Y%m%d')))}" 
        if code1 == code:
            return True
    return False




def set_styles():
    try:
        with open(racine/"style"/"style.qss", "r") as file:
            style = file.read()
            background_path_str =  str(background_path).replace("\\", "/") 
            arrowdrop_str =  str(arrowdrop).replace("\\", "/") 
            style = style.replace("background_image",background_path_str)
            style = style.replace("arrowdrop", arrowdrop_str) 
            return style
    except FileNotFoundError:
        print("Style file not found. Using default styles.")


def increment_month(year: int, month: int) -> tuple:
    """
    Incrémente un mois et ajuste l'année si nécessaire.

    Args:
        year (int): L'année actuelle.
        month (int): Le mois actuel (1 à 12).

    Returns:
        tuple: Une paire (année mise à jour, mois mis à jour).
    """
    if not (1 <= month <= 12):
        raise ValueError("Le mois doit être entre 1 et 12 inclus.")

    # Incrémenter le mois
    month += 1
    if month > 12:  # Si on dépasse décembre
        month = 1  # Revenir à janvier
        year += 1  # Incrémenter l'année

    return year, month


def decrement_month(year: int, month: int) -> tuple:
    """
    Décrémente un mois et ajuste l'année si nécessaire.

    Args:
        year (int): L'année actuelle.
        month (int): Le mois actuel (1 à 12).

    Returns:
        tuple: Une paire (année mise à jour, mois mis à jour).
    """
    if not (1 <= month <= 12):
        raise ValueError("Le mois doit être entre 1 et 12 inclus.")

    # Décrémenter le mois
    month -= 1
    if month < 1:  # Si on descend en dessous de janvier
        month = 12  # Revenir à décembre
        year -= 1  # Décrémenter l'année

    return year, month

def deplacer_et_renommer_image(source_path, destination_folder, nouveau_nom):
    """
    Déplace une image vers un dossier spécifique et la renomme.

    Args:
        source_path (str): Le chemin complet de l'image source.
        destination_folder (str): Le dossier de destination.
        nouveau_nom (str): Le nouveau nom du fichier (sans extension).
    
    Returns:
        str: Le nouveau chemin complet de l'image déplacée et renommée.

    """

    try:

        # Vérifier si le fichier source existe
        if not os.path.exists(source_path):
            raise FileNotFoundError(f"Le fichier source n'existe pas : {source_path}")
        
        # Créer le dossier de destination s'il n'existe pas
        os.makedirs(destination_folder, exist_ok=True)
        
        # Obtenir l'extension de l'image source
        extension = os.path.splitext(source_path)[1]
        
        # Construire le chemin complet du nouveau fichier
        nouveau_chemin = os.path.join(destination_folder, f"{nouveau_nom}{extension}")
        
        # Déplacer et renommer le fichier
        if source_path != nouveau_chemin:
            shutil.copy(source_path, nouveau_chemin)
        return nouveau_chemin
    except:
        
        return "Aucune"

    
    





def write_to_excel(file_name="output.xlsx"):
     
    workbook = Workbook()
    adherents = fetch_data()
    paiments = recuperer_all_paiements()
    depenses = recuperer_all_depenses1()
    salarie = recuperer_all_salarie()
    stock = recuperer_all_stock()
    ventes = recuperer_all_ventes()
    sanction = recuperer_all_sanction()



    # Feuille 1
    sheet1 = workbook.active
    sheet1.title = "adherents"
    for row in adherents:
        sheet1.append(row)

    # Feuille 2
    sheet2 = workbook.create_sheet(title="paiments")
    for row in paiments:
        sheet2.append(row)
    # Feuille 2
    sheet3 = workbook.create_sheet(title="depenses")
    for row in depenses:
        sheet3.append(row)

    # Sauvegarde du fichier Excel
    workbook.save(file_name) 






def generer_graphique_compta(donnees):
    mois = [d['mois'] for d in donnees]
    ventes = [d['ventes'] for d in donnees]
    achats = [d['achats'] for d in donnees]
    paiements = [d['paiments'] for d in donnees]
    depenses = [d['dépenses'] for d in donnees]
    salaires = [d['salaires'] for d in donnees]

    # Largeur des barres
    bar_width = 0.15
    x = np.arange(len(mois))  # Conversion pour gestion des calculs

    # Créer la figure et l'axe
    fig, ax = plt.subplots(figsize=(12, 6))
    fig.set_facecolor((0, 0, 0, 0.1))
    ax.set_facecolor((0, 0, 0, 0.1))

    # Tracer les différentes catégories
    ax.bar(x, paiements, bar_width, label='Paiements', color='#9ACD32', alpha=0.75)
    ax.bar(x, ventes, bar_width, bottom=paiements, label='Ventes', color='#98fb98', alpha=0.75)

    ax.bar(x - bar_width, achats, bar_width, label='Achats', color='#FF6347', alpha=0.75)
    ax.bar(x - bar_width, depenses, bar_width, bottom=achats, label='Dépenses', color='red', alpha=0.75)

    s = np.array(depenses) + np.array(achats)
    ax.bar(x - bar_width, salaires, bar_width, bottom=s, label='Salaires', color='#FF4500', alpha=0.75)

    # Ajouter les titres et étiquettes
    ax.set_title("Évolution des Comptes par Mois", color='white', fontsize=14)
    ax.set_xlabel("Mois", color='white', fontsize=12)
    ax.set_ylabel("Valeurs", color='white', fontsize=12)
    ax.set_xticks(x)
    ax.set_xticklabels(mois, rotation=90)
    ax.tick_params(axis='x', colors='white')  # Ticks de l'axe X en blanc
    ax.tick_params(axis='y', colors='white')

    # Légende
    ax.legend()

    # Ajuster les marges
    fig.subplots_adjust(left=0.2, right=0.9, top=0.9, bottom=0.25)
 

    image_stream = BytesIO()
    fig.savefig(image_stream, format='png')
    image_stream.seek(0)
    temp_image_file = NamedTemporaryFile(delete=False, suffix='.png')
    image_path = temp_image_file.name
    fig.savefig(image_path, format='png')
    temp_image_file.close()
    return image_path


    


def generate_compta_rapport(path): 
        

        data = recuperer_compta_each_month(24)
        data_year = recuperer_compta_each_year()
        
        
        path_evolution = generer_graphique_compta(data)        
        # Création d'un objet Canvas pour générer le PDF
        c = canvas.Canvas(path, pagesize=letter)

        # Ajout de texte dans le PDF
        c.setFont("Helvetica", 16)
        c.drawString(100, 700, f"Rapport de Comptabilité.")
        
        c.line(100, 680, 500, 680)
        c.setFont("Helvetica", 12)  
        c.drawString(100, 660, f"""Ce rapport donne un résumé de la comptabilité de l'année en cours""")
        c.drawString(100, 640, f"""et de l'année précédente.""")


        c.drawImage(path_evolution, 100, 360, width=450, height=250) 

        date_now = datetime.now().strftime("%Y-%m")
        

        revenue_ce_moi = data_year[-1]["ventes"] + data_year[-1]["paiements"] 
        depense_ce_moi = data_year[-1]["achats"] + data_year[-1]["dépenses"] + data_year[-1]["salaires"] 

        c.setFont("Helvetica", 14)
        # Ajouter du texte à différentes positions
        c.drawString(100, 340, f"Situation de l'année actuel {data_year[-1]["annee"]} : ")
        c.setFont("Helvetica", 12) 
        c.drawString(120, 320, f"Revenus par catégorie : ")
        c.drawString(130, 300, f"- Ventes {data_year[-1]["ventes"]} Dhs.")
        c.drawString(130, 280, f"- Paiments des adhérents :  {data_year[-1]["paiements"]} Dhs.")
        c.drawString(120, 260, f"Dépenses par catégorie :  ")
        c.drawString(130, 240, f"- Salaires {data_year[-1]["salaires"] } Dhs.")
        c.drawString(130, 220, f"- Achats  : {data_year[-1]["achats"]} Dhs.")
        c.drawString(130, 200, f"- Autres dépenses : {data_year[-1]["dépenses"]} Dhs.")
        c.drawString(120, 180, f"Résumé financier :")

        c.drawString(130, 160, f"- Revenus totaux :  {revenue_ce_moi} Dhs. ")
        c.drawString(130, 140, f"- Dépenses totales :  {depense_ce_moi} Dhs.")
        c.drawString(130, 120, f"- Résultat net :{revenue_ce_moi - depense_ce_moi} Dhs ")
        c.drawString(100, 100, f"La liste de chaque mois se trouve dans les pages suivantes.")


        data = recuperer_compta_each_month(24)

        for i in range(len(data)-1, 0, -2): 
            c.showPage()

            c.setFont("Helvetica", 14)
            # Ajouter du texte à différentes positions
            c.drawString(100, 700, f"Situation de mois : {data[i]["mois"]} : ")
            c.setFont("Helvetica", 12) 
            c.drawString(120, 680, f"Revenus par catégorie : ")
            c.drawString(130, 660, f"- Ventes {data[i]["ventes"]} Dhs.")
            c.drawString(130, 640, f"- Paiments des adhérents :  {data[i]["paiments"]} Dhs.")
            c.drawString(120, 620, f"Dépenses par catégorie :  ")
            c.drawString(130, 600, f"- Salaires {data[i]["salaires"] } Dhs.")
            c.drawString(130, 580, f"- Achats  : {data[i]["achats"]} Dhs.")
            c.drawString(130, 560, f"- Autres dépenses : {data[i]["dépenses"]} Dhs.")
            c.drawString(120, 540, f"Résumé financier :")
            c.drawString(130, 520, f"- Revenus totaux :  {revenue_ce_moi} Dhs. ")
            c.drawString(130, 500, f"- Dépenses totales :  {depense_ce_moi} Dhs.")
            c.drawString(130, 480, f"- Résultat net :{revenue_ce_moi - depense_ce_moi} Dhs ") 
            c.line(100, 460, 500, 460)
            c.setFont("Helvetica", 14)
            if i == 0:
                break
            # Ajouter du texte à différentes positions
            c.drawString(100, 440, f"Situation de mois : {data[i-1]["mois"]} : ")
            c.setFont("Helvetica", 12) 
            c.drawString(120, 420, f"Revenus par catégorie : ")
            c.drawString(130, 400, f"- Ventes {data[i-1]["ventes"]} Dhs.")
            c.drawString(130, 380, f"- Paiments des adhérents :  {data[i-1]["paiments"]} Dhs.")
            c.drawString(120, 360, f"Dépenses par catégorie :  ")
            c.drawString(130, 340, f"- Salaires {data[i-1]["salaires"] } Dhs.")
            c.drawString(130, 320, f"- Achats  : {data[i-1]["achats"]} Dhs.")
            c.drawString(130, 300, f"- Autres dépenses : {data[i-1]["dépenses"]} Dhs.")
            c.drawString(120, 280, f"Résumé financier :")
            c.drawString(130, 260, f"- Revenus totaux :  {revenue_ce_moi} Dhs. ")
            c.drawString(130, 240, f"- Dépenses totales :  {depense_ce_moi} Dhs.")
            c.drawString(130, 220, f"- Résultat net :{revenue_ce_moi - depense_ce_moi} Dhs ")

        c.save()


   